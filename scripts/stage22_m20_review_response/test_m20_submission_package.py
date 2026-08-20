#!/usr/bin/env python3
"""Structural, numerical, and reporting validation for the M20 upload set."""

from __future__ import annotations

import hashlib
import re
import zipfile
from pathlib import Path

import pandas as pd
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[2]
READY = ROOT / "submission_ready_scientific_reports_m20"
PACKAGE = ROOT / "submission_package" / "stage22_m20_review_response"
MANUSCRIPT_MD = ROOT / "manuscript_ready" / "stage22_m20_review_response" / "scientific_reports_m20_manuscript.md"
ROBUST = ROOT / "tables" / "stage22_m20_review_response"


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    expected = {
        "manuscript.docx", "cover_letter_AUTHOR_COMPLETION_REQUIRED.pdf", "PRISMA_2020_checklist.pdf",
        "supplementary_information.pdf", "Supplementary_Tables_S1-S28.xlsx",
        "Source_Data_M20.zip", "Source_Code_M20.zip",
        "Figure_1.pdf", "Figure_2.pdf", "Figure_3.pdf", "Figure_4.pdf",
    }
    require({p.name for p in READY.iterdir() if p.is_file()} == expected, "upload file set differs from the planned 11 files")
    require(all((READY / name).stat().st_size > 10_000 for name in expected), "unexpectedly small upload asset")

    text = MANUSCRIPT_MD.read_text(encoding="utf-8")
    abstract = text.split("## Abstract", 1)[1].split("## Keywords", 1)[0]
    words = re.findall(r"[A-Za-z0-9]+(?:[–-][A-Za-z0-9]+)*", abstract)
    require(len(words) <= 200, f"abstract exceeds 200 words: {len(words)}")
    required_phrases = [
        "not prospectively preregistered", "common measurable intersections",
        "100,000 Monte Carlo allocations", "studentized maxT",
        "bootstrap 95% CI", "0/783", "not falsely presented as an independent human dual screen",
        "Absence of a located statement is a provenance gap, not an allegation",
        "GSE99339 contains the same archived H7 glomerular donor labels",
        "AI did not autonomously determine which inferential conclusions to retain",
    ]
    for phrase in required_phrases:
        require(phrase in text, f"missing required manuscript disclosure: {phrase}")
    require("vascular-wall interaction" in abstract and "coagulation met it in zero sources" in abstract, "abstract does not reflect M20 results")
    refs = [int(x) for x in re.findall(r"^(\d+)\.", text, flags=re.MULTILINE)]
    require(refs == list(range(1, 31)), f"reference list is not sequential 1–30: {refs}")
    require("10.1038/s41598-025-01628-5" in text and "10.1111/jdi.13739" in text, "closest prior studies are not cited")
    require(all(f"GSE{a}" in text for a in [1009, 30528, 30529, 96804, 104948, 104954, 111154, 142025, 163603, 166239, 199838]), "formal GEO citations incomplete")

    doc = Document(READY / "manuscript.docx")
    doc_text = "\n".join(p.text for p in doc.paragraphs)
    require("Common-measurement pathway associations" in doc_text, "DOCX missing revised Figure 4 legend")
    require("Source_Code_M19" not in doc_text and "v1.1.2" not in doc_text, "stale M19 availability text in DOCX")

    checklist_text = "\n".join(page.extract_text() or "" for page in PdfReader(READY / "PRISMA_2020_checklist.pdf").pages)
    for subitem in ["10b", "13b", "13c", "13d", "13e", "13f", "16b", "20a", "20b", "20c", "20d", "23d", "24c"]:
        require(subitem in checklist_text, f"PRISMA subitem absent: {subitem}")

    cover_text = "\n".join(page.extract_text() or "" for page in PdfReader(READY / "cover_letter_AUTHOR_COMPLETION_REQUIRED.pdf").pages)
    cover_text = re.sub(r"\s+", " ", cover_text)
    for field in ["Suggested reviewers", "Referees to exclude", "Editorial Board Member", "AUTHOR TO COMPLETE", "AUTHOR TO CONFIRM"]:
        require(field in cover_text, f"author-only cover-letter field missing: {field}")
    require("v1.2.1" in cover_text, "cover letter does not cite the M20.1 archive")

    for index in range(1, 5):
        reader = PdfReader(READY / f"Figure_{index}.pdf")
        require(len(reader.pages) == 1, f"Figure {index} must be one page")
        page = reader.pages[0]
        require(float(page.mediabox.width) > 400 and float(page.mediabox.height) > 250, f"Figure {index} page too small")

    wb = load_workbook(READY / "Supplementary_Tables_S1-S28.xlsx", read_only=True)
    expected_sheets = {"README", "Data_dictionary"} | {f"Table_S{i}" for i in range(1, 29)}
    require(set(wb.sheetnames) == expected_sheets, "supplement workbook sheet set incomplete")
    overlap_rows = list(wb["Table_S28"].iter_rows(values_only=True))
    overlap_text = "\n".join("|".join(str(value or "") for value in row) for row in overlap_rows)
    for donor in ["DN901", "DN910", "DN914", "DN916", "DN932", "DN941", "DN947"]:
        require(donor in overlap_text, f"source-lineage audit missing repeated donor {donor}")
    require("GSE47183" in overlap_text and "GSE32591" in overlap_text, "historical CDF lineage missing")

    with zipfile.ZipFile(READY / "Source_Code_M20.zip") as archive:
        names = set(archive.namelist())
        required_code = {
            "README_REPRODUCE.md", "requirements-m20.txt",
            "scripts/stage22_m20_review_response/run_m20_robustness.py",
            "scripts/stage22_m20_review_response/test_m20_reproducibility.py",
            "data_processed/m20_primary_reproduction/input_manifest.csv",
            "data_processed/m20_primary_reproduction/canonical_pathways.gmt",
        }
        require(required_code <= names, f"code archive missing: {required_code - names}")
        require(any(name.endswith("__canonical_expression.csv.gz") for name in names), "frozen expression inputs missing from code archive")
    with zipfile.ZipFile(READY / "Source_Data_M20.zip") as archive:
        names = set(archive.namelist())
        require(all(f"Supplementary_Table_S{i}.csv" in names for i in range(1, 29)), "source-data archive missing supplementary table")
        require(any(name.startswith("systematic_search/") for name in names), "search evidence missing")

    manifest = pd.read_csv(PACKAGE / "final_upload_manifest.csv")
    require(set(manifest["upload_file"]) == expected, "upload manifest filenames differ")
    for row in manifest.itertuples(index=False):
        require(sha256(READY / row.upload_file) == row.sha256, f"upload hash mismatch: {row.upload_file}")

    rep = pd.read_csv(ROBUST / "primary_pathway_operational_replication.csv")
    require(int(rep["operational_two_of_three_call"].sum()) == 4, "expected four operational pathway calls")
    coverage = pd.read_csv(ROBUST / "primary_pathway_measurement_coverage.csv")
    require(int(coverage.loc[coverage["reactome_name"].eq("Complement cascade"), "common_three_cohort_members"].iloc[0]) == 51, "complement common coverage changed")

    print(f"M20_SUBMISSION_VALIDATION=PASS abstract_words={len(words)} ready_files={len(expected)} prisma_rows=42")


if __name__ == "__main__":
    main()
