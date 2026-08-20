#!/usr/bin/env python3
"""Contract tests for the M19 submission-ready package."""

from __future__ import annotations

import hashlib
import re
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


ROOT=Path(__file__).resolve().parents[2]
TABLES=ROOT/"tables"/"stage21_m19_scientific_reports_revision"
MANUSCRIPT=ROOT/"manuscript_ready"/"stage21_m19_scientific_reports_revision"/"scientific_reports_m19_manuscript.md"
PACKAGE=ROOT/"submission_package"/"stage21_m19_scientific_reports_revision"
SUPP=PACKAGE/"supplementary"
READY=ROOT/"submission_ready_scientific_reports_m19"
EXPECTED={"manuscript.docx","cover_letter.pdf","PRISMA_2020_checklist.pdf","supplementary_information.pdf","Supplementary_Tables_S1-S19.xlsx","Source_Data_M19.zip","Source_Code_M19.zip",*{f"Figure_{i}.png" for i in range(1,5)}}


def sha256(path:Path)->str:
    h=hashlib.sha256(); h.update(path.read_bytes()); return h.hexdigest()


def main()->None:
    assert {p.name for p in READY.iterdir() if p.is_file()}==EXPECTED
    text=MANUSCRIPT.read_text(encoding="utf-8")
    title=text.splitlines()[0].removeprefix("# "); assert len(title.split())<=20
    abstract=text.split("## Abstract",1)[1].split("## Keywords",1)[0]
    abstract_words=re.findall(r"\b[\w–-]+\b",abstract); assert len(abstract_words)<=200,len(abstract_words)
    assert len(text.split("## Keywords",1)[1].split("## Introduction",1)[0].strip().split(";"))==6
    assert "0/783" in text and "coagulation pathway did not replicate" in text.lower()
    assert not re.search(r"(?:P|p)\s*≤",text)
    screen=pd.read_csv(TABLES/"systematic_dataset_screening_m19.csv")
    assert len(screen)==322
    assert screen["final_decision"].value_counts().to_dict()=={"EXCLUDED_TITLE_SUMMARY":269,"EXCLUDED_FULL_RECORD":42,"INCLUDED":11}
    cohort=pd.read_csv(TABLES/"cohort_characteristics.csv"); assert len(cohort)==11 and cohort["source_study"].nunique()==9
    meta=pd.read_csv(TABLES/"primary_glomerular_gene_meta.csv"); assert len(meta)==783 and int(meta["fdr_lt_0_05"].sum())==0 and int(meta["k"].eq(3).sum())==582
    rep=pd.read_csv(TABLES/"primary_glomerular_pathway_replication_summary.csv")
    called=set(rep.loc[rep["primary_replication_call"],"reactome_name"])
    assert called=={"Complement cascade","Chemokine receptors bind chemokines","Extracellular matrix organization"}
    assert not bool(rep.loc[rep["reactome_name"].eq("Coagulation pathway"),"primary_replication_call"].iloc[0])
    for path in SUPP.glob("Supplementary_Table_S*.csv"):
        frame=pd.read_csv(path,nrows=100)
        for col in frame.columns: assert not frame[col].astype(str).str.contains(r"^[A-Za-z]:\\",regex=True).any(),(path,col)
    book=load_workbook(READY/"Supplementary_Tables_S1-S19.xlsx",read_only=True,data_only=True)
    assert set(book.sheetnames)=={"README","Data_dictionary",*{f"Table_S{i}" for i in range(1,20)}}; book.close()
    for i in range(1,5):
        with Image.open(READY/f"Figure_{i}.png") as im: assert im.width>=2000 and im.height>=1500
    for name in ("cover_letter.pdf","PRISMA_2020_checklist.pdf","supplementary_information.pdf"):
        assert len(PdfReader(READY/name).pages)>=1
    with zipfile.ZipFile(READY/"manuscript.docx") as z:
        xml="\n".join(z.read(n).decode("utf-8",errors="ignore") for n in z.namelist() if n.startswith("word/") and n.endswith(".xml"))
        assert "w:lnNumType" in xml and " PAGE " in xml
    with zipfile.ZipFile(READY/"Source_Code_M19.zip") as z:
        names=set(z.namelist())
        for script in ("run_systematic_geo_search.py","run_systematic_arrayexpress_search.py","run_systematic_pubmed_search.py","run_m19_compartment_analysis.py","build_final_screening_table.py","build_m19_submission_package.py"):
            assert f"scripts/stage21_m19_scientific_reports_revision/{script}" in names
    manifest=pd.read_csv(PACKAGE/"final_upload_manifest.csv")
    for row in manifest.itertuples(): assert sha256(READY/row.upload_file)==row.sha256
    print(f"M19_SUBMISSION_VALIDATION=PASS abstract_words={len(abstract_words)} ready_files={len(EXPECTED)}")


if __name__=="__main__": main()
